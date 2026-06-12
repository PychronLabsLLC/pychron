"""Verify xlsx_exporter output against the HW reference uploads.

Run:
    python -m pychron.ausgeochem.tests.xlsx_export_test
"""

import datetime
import os
import sys
import tempfile

from openpyxl import load_workbook
from uncertainties import ufloat

from pychron.ausgeochem.xlsx_exporter import (
    API_FIELD_ROW,
    EarthBankXlsxExporter,
)

HW_BASE = (
    "/Users/jross/Library/CloudStorage/GoogleDrive-Jake.Ross@nmt.edu/"
    "My Drive/ 0.1_Pychron/ PychronConsulting/EarthBank"
)
HW_ARAR = os.path.join(HW_BASE, "ArArDatapoint.template.v2025-11-05_HW1-2Only_Test.xlsx")
HW_SAMPLE = os.path.join(HW_BASE, "HWSampleUpload.xlsx")

ARAR_SHEET_MAP = {
    "ArArDataPoints": "ArAr Datapoints",
    "ArArMeasurement": "Measurements",
    "ArArAliquot": "Aliquots",
    "ArArAgeSummary": "Age Summaries",
    "ArArAgeCalc": "Age Calculations",
}


# -- fake pychron Analysis / AnalysisGroup -----------------------------------


class _FakeIso(object):
    def __init__(self, v):
        self.v = v

    def get_interference_corrected_value(self):
        return ufloat(self.v, self.v * 0.01)

    def get_non_detector_corrected_value(self):
        return ufloat(self.v * 1.1, self.v * 0.02)

    class _Blank(object):
        @property
        def uvalue(self):
            return ufloat(0.01, 0.001)

    blank = _Blank()


class _FakeAnalysis(object):
    labnumber = "HW1"
    aliquot = 1
    step = ""
    rundate = datetime.datetime(2024, 3, 5, 12, 30)
    analysis_timestamp = datetime.datetime(2024, 3, 5, 12, 30, 15)
    timestamp = None
    weight = 2.5
    beam_diameter = 70
    extract_value = 1200
    extract_units = "C"
    radiogenic_yield = 98.5
    is_plateau_step = True
    sample = "HW-001"
    project = "EarthBank-Test"
    experiment_type = "Ar/Ar"
    analysis_type = "unknown"
    material = "sanidine"
    lithology = "granite"
    grainsize = "75-150"
    latitude = -34.93
    longitude = 138.6
    elevation = 100.0
    j = ufloat(1e-3, 1e-6)
    rad40 = ufloat(9.8, 0.1)
    k39 = ufloat(1.0, 0.01)
    uage = ufloat(28.2, 0.05)
    cak = ufloat(0.1, 0.001)
    kca = ufloat(10, 0.1)
    clk = ufloat(0.01, 0.0001)
    k2o = ufloat(8, 0.1)
    sample_note = "fake test analysis"
    sample_prep_comment = ""

    def get_isotope(self, n):
        return {
            "Ar40": _FakeIso(10),
            "Ar39": _FakeIso(1),
            "Ar38": _FakeIso(0.1),
            "Ar37": _FakeIso(0.05),
            "Ar36": _FakeIso(0.001),
        }.get(n)

    def get_ratio(self, k):
        return ufloat(0.5, 0.01)


class _FakeAnalysis2(_FakeAnalysis):
    aliquot = 2
    step = ""
    is_plateau_step = True


class _FakeAG(object):
    age_error_kind = "1 sigma"
    grainsize = "75-150"
    lithology = "granite"
    material = "sanidine"
    sample = "HW-001"
    project = "EarthBank-Test"
    monitor_info = "FCs"
    comments = "Verification group"

    class _C:
        age_units = "Ma"
        lambda_b_citation = "Min (2008)"
        lambda_e_citation = "Min (2008)"
        atm4036_citation = "Lee (2006)"

    arar_constants = _C()

    def __init__(self):
        self.analyses = [_FakeAnalysis(), _FakeAnalysis2()]
        self.nanalyses = 2

    def get_preferred_obj(self, k):
        class _P:
            kind = "Plateau Age"
            computed_kind = "Plateau Age"
            error_kind = "1 sigma"

        return _P()

    def get_preferred_mswd_tuple(self):
        return (1.05, True, 2, 0.42)

    def get_ma_scaled_age(self):
        return ufloat(28.2, 0.05)


# -- verification ------------------------------------------------------------


def _api_fields(path, sheet_name):
    wb = load_workbook(path, data_only=False)
    sheet = wb[sheet_name]
    keys = []
    for col in range(1, sheet.max_column + 1):
        v = sheet.cell(row=API_FIELD_ROW, column=col).value
        if v is not None and str(v).strip():
            keys.append(str(v).strip())
    return keys


def _data_rows(path, sheet_name, start=5):
    wb = load_workbook(path, data_only=False)
    sheet = wb[sheet_name]
    rows = []
    for r in range(start, sheet.max_row + 1):
        row = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
        if any(v not in (None, "") for v in row):
            rows.append(row)
    return rows


def main():
    out_dir = tempfile.mkdtemp(prefix="earthbank_xlsx_")
    print("output dir:", out_dir)

    ag = _FakeAG()
    exp = EarthBankXlsxExporter()
    arar_out = os.path.join(out_dir, "ArArDataPoint_HW.xlsx")
    sample_out = os.path.join(out_dir, "Sample_HW.xlsx")
    exp.export_analysis_group(ag, arar_out, datapoint_key="HW-001")
    exp.export_sample(ag, sample_out)

    print("\n=== ArArDataPoint columns vs HW ===")
    arar_failures = 0
    for tmpl_name, hw_name in ARAR_SHEET_MAP.items():
        out_keys = _api_fields(arar_out, tmpl_name)
        hw_keys = _api_fields(HW_ARAR, hw_name)
        common = set(out_keys) & set(hw_keys)
        only_out = set(out_keys) - set(hw_keys)
        only_hw = set(hw_keys) - set(out_keys)
        # HW columns must be a subset (HW shouldn't reference fields the export lacks)
        # but some renames (datapoint.funding vs funding) are expected.
        status = "OK" if not only_hw or only_hw <= {"funding", "literature"} else "DIFF"
        if status != "OK":
            arar_failures += 1
        print(
            "  [{:4s}] {:18s} out={:>2d}  hw={:>2d}  common={:>2d}  hw_only={}".format(
                status,
                tmpl_name,
                len(out_keys),
                len(hw_keys),
                len(common),
                sorted(only_hw)[:5],
            )
        )

    print("\n=== ArArDataPoint row counts (exported) ===")
    for tmpl_name in ARAR_SHEET_MAP:
        n = len(_data_rows(arar_out, tmpl_name))
        print("  {:18s} -> {} data row(s)".format(tmpl_name, n))

    print("\n=== Sample columns vs HW ===")
    out_keys = _api_fields(sample_out, "Samples")
    hw_keys = _api_fields(HW_SAMPLE, "Samples")
    common = set(out_keys) & set(hw_keys)
    only_hw = set(hw_keys) - set(out_keys)
    print(
        "  out={} hw={} common={} hw_only={}".format(
            len(out_keys), len(hw_keys), len(common), sorted(only_hw)[:8]
        )
    )
    sample_rows = _data_rows(sample_out, "Samples")
    print("  exported Sample rows: {}".format(len(sample_rows)))

    print("\noutputs:")
    print("  arar:   {}".format(arar_out))
    print("  sample: {}".format(sample_out))

    sys.exit(1 if arar_failures else 0)


if __name__ == "__main__":
    main()
