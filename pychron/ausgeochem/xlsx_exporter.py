# ===============================================================================
# Copyright 2024 Pychron Developers
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ===============================================================================

"""Render EarthBank upload payloads to the official AusGeochem xlsx templates.

The EarthBank web UI accepts the same five-sheet ArArDataPoint workbook and
single-sheet Samples workbook that this exporter produces, so the output is
interchangeable with a manual template upload. Useful for dry-runs, archival,
and offline review.
"""

from __future__ import absolute_import

import os
import shutil

from openpyxl import load_workbook

from pychron.ausgeochem.earthbank_service import AusGeochemEarthBankService


TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "templates")
ARAR_TEMPLATE = os.path.join(TEMPLATE_DIR, "ArArDataPoint_Template2026.xlsx")
SAMPLE_TEMPLATE = os.path.join(TEMPLATE_DIR, "Sample_Template_v2025-04-16.xlsx")

# Header rows 0..3 = description / type / human label / apiField.
# Data writes start at row index 4 (i.e. spreadsheet row 5).
API_FIELD_ROW = 4  # openpyxl is 1-indexed
DATA_START_ROW = 5

# Sheets that don't follow the rich 4-header-rows layout. Map: sheet_name ->
# (apiField_row, data_start_row). Default is (API_FIELD_ROW, DATA_START_ROW).
SHEET_LAYOUT = {
    "Datapoint Props": (1, 2),
    "Sample Props": (1, 2),
}

# Per-sheet name → service payload key overrides. Used when the spreadsheet
# column key does not match the DTO field by direct or `Name`-suffix lookup.
_MISSING = object()

SHEET_OVERRIDES = {
    "ArArDataPoints": {
        "arArSetUp": "arArAnalyticalSetUpId",
        "arArMethod": "arMethodName",
        "irradiationBatchID": "irradiationBatchIDId",
        "datapoint.literature": "literatureName",
        "datapoint.funding": "fundingName",
    },
    "ArArMeasurement": {
        "jvalueUncertaintyType": "jvalueUncertaintyTypeName",
        "jvalueUncertaintyUnit": "jvalueUncertaintyUnitName",
        "apparentAgeUncertainityType": "apparentAgeUncertainityTypeName",
    },
    "Samples": {
        "sampleName": "name",
        "sampleComment": "description",
        "lastKnownArchive": "archiveName",
        "elevationGround": "referenceElevation",
        "elevationDepthMin": "relativeElevationMin",
        "elevationDepthMax": "relativeElevationMax",
        "elevationDepthAccuracy": "relativeElevationAccuracy",
        "ageMin": "rockUnitAgeMin",
        "ageMax": "rockUnitAgeMax",
        "unitAgeDescription": "rockUnitAgeDescription",
        "unitName": "stratographicUnitName",
        "funding": "fundingName",
        "literature": "literatureName",
        "Tag": None,
    },
}


def _read_api_fields(sheet, header_row=API_FIELD_ROW):
    """Return a list of (col_index, apiField_key) for the apiField header row."""
    keys = []
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=header_row, column=col).value
        if val is None or val == "":
            continue
        keys.append((col, str(val).strip()))
    return keys


def _resolve_value(sheet_name, col_key, payload):
    overrides = SHEET_OVERRIDES.get(sheet_name, {})
    if col_key in overrides:
        ov = overrides[col_key]
        if ov is None:
            return None
        return payload.get(ov)
    if col_key in payload:
        return payload[col_key]
    if (col_key + "Name") in payload:
        return payload[col_key + "Name"]
    return None


def _write_row(sheet, row_idx, payload, col_keys, fk_value=None, fk_col="datapointName"):
    for col_idx, key in col_keys:
        if key == fk_col and fk_value is not None:
            sheet.cell(row=row_idx, column=col_idx).value = fk_value
            continue
        val = _resolve_value(sheet.title, key, payload)
        if val is None:
            continue
        sheet.cell(row=row_idx, column=col_idx).value = val


class EarthBankXlsxExporter(object):
    """Build AusGeochem-template xlsx files from pychron analysis groups.

    A standalone tool — instantiates a service in unbound mode purely so it
    can reuse the existing payload builders. No network calls are made.
    """

    def __init__(self, service=None):
        self._svc = service or AusGeochemEarthBankService(bind=False)

    def export_analysis_group(self, analysis_group, output_path, datapoint_key=None):
        """Render a single AnalysisGroup as an ArArDataPoint workbook.
        Convenience wrapper around :meth:`export_analysis_groups`."""

        return self.export_analysis_groups(
            [(analysis_group, datapoint_key)], output_path
        )

    def export_analysis_groups(self, items, output_path):
        """Render N AnalysisGroups into a single ArArDataPoint workbook.

        ``items`` is an iterable of ``(analysis_group, datapoint_key)`` tuples
        or bare AnalysisGroups. Each group becomes one ArArDataPoint /
        ArArAgeSummary / ArArAgeCalc row, one ArArAliquot row per unique
        aliquot, and one ArArMeasurement row per analysis — all keyed by a
        distinct ``datapointName``.
        """

        # Normalize input + resolve datapoint keys, ensuring uniqueness
        groups = []
        seen_keys = {}
        for entry in items:
            if isinstance(entry, tuple):
                ag, key = entry
            else:
                ag, key = entry, None
            analyses = list(getattr(ag, "analyses", None) or [])
            if not analyses:
                continue
            if not key:
                key = (
                    getattr(analyses[0], "sample", None)
                    or getattr(ag, "sample", None)
                    or "datapoint"
                )
            base = key
            i = 2
            while key in seen_keys:
                key = "{}-{}".format(base, i)
                i += 1
            seen_keys[key] = True
            groups.append((ag, analyses, key))

        if not groups:
            raise ValueError("No AnalysisGroups with analyses to export")

        shutil.copyfile(ARAR_TEMPLATE, output_path)
        wb = load_workbook(output_path)

        for ag, analyses, key in groups:
            dp_payload = self._svc.build_data_point_payload(
                analysis_group=ag, analysis=analyses[0]
            )
            self._append_rows(
                wb, "ArArDataPoints", [dp_payload], fk_value=key,
                fk_col="datapointName",
            )

            aliquot_rows = []
            seen = set()
            for a in analyses:
                aname = self._svc._aliquot_name(a)
                if aname in seen:
                    continue
                seen.add(aname)
                ap = self._svc.build_aliquot_payload(analysis=a)
                ap["aliquotName"] = aname
                aliquot_rows.append(ap)
            self._append_rows(
                wb, "ArArAliquot", aliquot_rows, fk_value=key,
                fk_col="datapointName",
            )

            meas_rows = [
                self._svc.build_measurement_payload(analysis=a) for a in analyses
            ]
            self._append_rows(
                wb, "ArArMeasurement", meas_rows, fk_value=key,
                fk_col="datapointName",
            )

            self._append_rows(
                wb, "ArArAgeSummary",
                [self._svc.build_age_summary_payload(ag)],
                fk_value=key, fk_col="datapointName",
            )
            self._append_rows(
                wb, "ArArAgeCalc",
                [self._svc.build_age_calc_payload(ag)],
                fk_value=key, fk_col="datapointName",
            )

            # Datapoint Props — emit pychron-specific metadata
            props = self._svc.build_datapoint_props(ag)
            if props and "Datapoint Props" in wb.sheetnames:
                self._append_rows(
                    wb, "Datapoint Props", props, fk_value=key,
                    fk_col="datapointName",
                )

        wb.save(output_path)
        return output_path

    def export_samples(self, analysis_groups, output_path):
        """Render N samples into a single Sample workbook (one row each).
        Duplicate sampleName entries are de-duplicated."""

        rows = []
        seen = set()
        for ag in analysis_groups:
            analyses = list(getattr(ag, "analyses", None) or [])
            if not analyses:
                continue
            sample_dto, location_dto = self._svc.build_sample_payload(
                analysis=analyses[0], analysis_group=ag
            )
            sname = sample_dto.get("name")
            if sname in seen:
                continue
            seen.add(sname)
            merged = dict(sample_dto)
            for k in ("lat", "lon", "latLonPrecision"):
                if k in location_dto:
                    target = {
                        "lat": "latitude", "lon": "longitude",
                        "latLonPrecision": "latLonPrecision",
                    }[k]
                    merged.setdefault(target, location_dto[k])
            rows.append(merged)

        if not rows:
            raise ValueError("No samples to export")

        shutil.copyfile(SAMPLE_TEMPLATE, output_path)
        wb = load_workbook(output_path)
        self._append_rows(wb, "Samples", rows)

        # Sample Props per row
        if "Sample Props" in wb.sheetnames:
            for ag in analysis_groups:
                analyses = list(getattr(ag, "analyses", None) or [])
                if not analyses:
                    continue
                a = analyses[0]
                sample_name = getattr(a, "sample", None) or getattr(ag, "sample", None)
                if not sample_name:
                    continue
                props = self._svc.build_sample_props(analysis=a, analysis_group=ag)
                if props:
                    self._append_rows(
                        wb, "Sample Props", props, fk_value=sample_name,
                        fk_col="sampleName",
                    )
        wb.save(output_path)
        return output_path

    def export_sample(self, analysis_group, output_path, analysis=None):
        """Render a Sample workbook for the analysis group's sample."""

        analyses = list(getattr(analysis_group, "analyses", None) or [])
        if not analyses and analysis is None:
            raise ValueError("AnalysisGroup has no analyses and no analysis arg")
        if analysis is None:
            analysis = analyses[0]

        sample_dto, location_dto = self._svc.build_sample_payload(
            analysis=analysis, analysis_group=analysis_group
        )
        # Flatten location into sample for the spreadsheet's wide row layout.
        merged = dict(sample_dto)
        for k in ("lat", "lon", "latLonPrecision"):
            if k in location_dto:
                merged.setdefault(
                    {"lat": "latitude", "lon": "longitude",
                     "latLonPrecision": "latLonPrecision"}[k],
                    location_dto[k],
                )

        shutil.copyfile(SAMPLE_TEMPLATE, output_path)
        wb = load_workbook(output_path)
        self._append_rows(wb, "Samples", [merged])

        # Sample Props
        sample_name = merged.get("name") or merged.get("sampleID")
        props = self._svc.build_sample_props(
            analysis=analysis, analysis_group=analysis_group
        )
        if props and sample_name and "Sample Props" in wb.sheetnames:
            self._append_rows(
                wb, "Sample Props", props, fk_value=sample_name,
                fk_col="sampleName",
            )
        wb.save(output_path)
        return output_path

    # ------------------------------------------------------------------
    def _append_rows(self, wb, sheet_name, payloads, fk_value=None,
                     fk_col="datapointName"):
        if sheet_name not in wb.sheetnames:
            raise KeyError("template missing sheet {!r}".format(sheet_name))
        sheet = wb[sheet_name]
        header_row, data_start = SHEET_LAYOUT.get(
            sheet_name, (API_FIELD_ROW, DATA_START_ROW)
        )
        col_keys = _read_api_fields(sheet, header_row)
        # Find first empty row at or below data_start
        first = data_start
        while sheet.cell(row=first, column=1).value not in (None, ""):
            first += 1
        for offset, payload in enumerate(payloads):
            _write_row(
                sheet, first + offset, payload, col_keys,
                fk_value=fk_value, fk_col=fk_col,
            )


# ============= EOF =============================================
